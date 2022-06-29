from openpyxl import workbook, load_workbook
from getting_report_data import *


def creation_report_xlsx():
    wb = load_workbook('СИР ГБУ свыше 5 мес..xlsx')
    ws = wb.active
    data = get_report_data()
    
    l_cl = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'T', 'U']
    start_y = 5
    x_now = 0
    for cellObject in data:
        x_now = 0
        for cell1 in data[cellObject]:
            ws[l_cl[x_now]+str(start_y)].value = data[cellObject][cell1]
            x_now += 1
        start_y += 1
    wb.save('Отчет.xlsx')
