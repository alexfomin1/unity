from command_sender import *
from command_sender_one import *
from command_sender_all import *
from districts_list import fl_317
from openpyxl import workbook, load_workbook

def get_report_data():
    start_y = 4
    
    data = {}
    
    districts = ['fl_317']
    # , 'fl_311', 'fl_312', 'fl_313', 'fl_314', 'fl_315', 'fl_316', 'fl_318', 'fl_319', 'fl_320', 'fl_321', 'fl_322', 'fl_323'
    
    for x in districts:
        data[x] = {}
        start_x = 2
        start_y += 1
        command_add = ('''
        SELECT COUNT(fls)
        FROM {}
        ''').format(x+'_c')
        data[x]['fls'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(summa)
        FROM {}
        ''').format(x+'_c')
        data[x]['summa'] = str(sendDataOne(command_add)[0])
        
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_sud')
        data[x]['fls_sud'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(summa_sud)
        FROM {}
        ''').format(x+'_sud')
        data[x]['summa_sud'] = str(sendDataOne(command_add)[0])
        
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_ssp')
        data[x]['fls_ssp'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(remains_ssp)
        FROM {}
        ''').format(x+'_ssp')
        data[x]['remains_ssp'] = str(sendDataOne(command_add)[0])
    
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_bank')
        data[x]['fls_bank'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(remains_bank)
        FROM {}
        ''').format(x+'_bank')
        data[x]['remains_bank'] = str(sendDataOne(command_add)[0])
        
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_restruct')
        data[x]['fls_restruct'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(remains_restruct)
        FROM {}
        ''').format(x+'_restruct')
        data[x]['remains_restruct'] = str(sendDataOne(command_add)[0])
        
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_impossible')
        data[x]['fls_impossible'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(summa_impossible)
        FROM {}
        ''').format(x+'_impossible')
        data[x]['summa_impossible'] = str(sendDataOne(command_add)[0])
        
        command_add = ('''
        SELECT COUNT(DISTINCT fls)
        FROM {}
        ''').format(x+'_no')
        data[x]['fls_no'] = str(sendDataOne(command_add)[0])
        command_add = ('''
        SELECT SUM(summa_more5)
        FROM {}
        ''').format(x+'_no')
        data[x]['summa_no'] = str(sendDataOne(command_add)[0])
        '''
        ws['C'+str(start_y)].value = data[x]['fls']
        ws['D'+str(start_y)].value = data[x]['summa']
        ws['E'+str(start_y)].value = data[x]['fls_sud']
        ws['F'+str(start_y)].value = data[x]['summa_sud']
        ws['G'+str(start_y)].value = data[x]['fls_ssp']
        ws['H'+str(start_y)].value = data[x]['remains_ssp']
        ws['I'+str(start_y)].value = data[x]['fls_bank']
        ws['K'+str(start_y)].value = data[x]['remains_bank']
        ws['L'+str(start_y)].value = data[x]['fls_restruct']
        ws['M'+str(start_y)].value = data[x]['remains_restruct']
        ws['N'+str(start_y)].value = data[x]['fls_impossible']
        ws['R'+str(start_y)].value = data[x]['summa_impossible']
        ws['T'+str(start_y)].value = data[x]['fls_no']
        ws['U'+str(start_y)].value = data[x]['summa_no']
     
        start_y += 1
        '''
    return data